"""
PriceCatalog — baza cen jednostkowych pozycji przedmiaru, edytowalna w Excelu
(data/price_catalog.xlsx, arkusz "Ceny").

Kluczem pozycji jest jej "opis" — dokładnie ten sam tekst, który
TakeoffCalculator generuje dla danej pozycji przedmiaru (np.
"Słup prefabrykowany — materiał"). Po tym tekście program dopasowuje cenę.

Samorozbudowa katalogu:
Przy każdym liczeniu przedmiaru (sync_and_price_items) lista pozycji zwrócona
przez TakeoffCalculator jest porównywana z zawartością arkusza "Ceny". Każda
pozycja, której opisu jeszcze w arkuszu nie ma, zostaje dopisana jako nowy
wiersz z pustą ceną. Dzięki temu w miarę rozbudowy programu (nowe fabryki /
nowe rodzaje elementów -> nowe pozycje w TakeoffCalculator) katalog sam
uzupełnia się o brakujące wiersze, gotowe do ręcznego wpisania ceny w Excelu —
nie trzeba nigdzie osobno "rejestrować" nowej pozycji cenowej.

Model edycji — POBIERZ / EDYTUJ LOKALNIE / WGRAJ:
Backend może być uruchomiony na innym komputerze niż przeglądarka (np. na
serwerze, z wieloma użytkownikami łączącymi się przez przeglądarkę). Dlatego
plik NIE jest otwierany bezpośrednio na serwerze — użytkownik pobiera go na
swój komputer (get_catalog_bytes), edytuje we własnym Excelu, po czym wgrywa
zaktualizowaną wersję z powrotem (import_price_catalog). Dzięki temu serwerowy
plik nigdy nie jest jednocześnie otwarty w Excelu i modyfikowany przez backend
— znika ryzyko konfliktu "naruszenia zasad współużytkowania" charakterystyczne
dla trzymania jednego pliku otwartego równolegle przez dwa programy.

Wgranie pliku SCALA go z aktualnym stanem na serwerze (a nie zastępuje w
ciemno) — pozycje dopisane automatycznie po stronie serwera pomiędzy
pobraniem a wgraniem (np. przez innego użytkownika liczącego przedmiar w tym
czasie) nie giną.

Wszystkie zapisy do pliku katalogu są serializowane (_LOCK) i atomowe (zapis
do pliku tymczasowego + podmiana), żeby dwa równoległe żądania nigdy nie
nadpisały się nawzajem ani nie zostawiły pliku w połowicznym stanie.
"""

from __future__ import annotations

import io
import os
import tempfile
import threading
import time
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_DATA_DIR = os.path.join(os.path.dirname(__file__), "..", "data")
CATALOG_PATH = os.path.join(_DATA_DIR, "price_catalog.xlsx")

PRICES_SHEET = "Ceny"
COL_OPIS = "Opis pozycji"
COL_JEDNOSTKA = "Jednostka"
COL_CENA = "Cena jednostkowa [PLN]"
COL_UWAGI = "Uwagi"
PRICE_COLUMNS = [COL_OPIS, COL_JEDNOSTKA, COL_CENA, COL_UWAGI]

INSTRUCTION_LINES = [
    "INSTRUKCJA EDYCJI KATALOGU CEN JEDNOSTKOWYCH",
    "",
    "Ten plik zasila kolumny 'Cena jednostkowa' i 'Wartość' w przedmiarze ilościowym",
    "(zakładka 'Przedmiar' w aplikacji oraz eksport do Excela).",
    "",
    "Zasady:",
    "1. Nie zmieniaj nazwy arkusza 'Ceny' ani nazw kolumn w wierszu 1.",
    "2. Kolumna 'Opis pozycji' MUSI dokładnie odpowiadać opisowi generowanemu przez",
    "   program (wielkość liter, myślniki, spacje) — to po niej program dopasowuje cenę.",
    "   Nie edytuj tego tekstu ręcznie.",
    "3. Wpisz cenę w kolumnie 'Cena jednostkowa [PLN]' dla pozycji, którą chcesz wycenić.",
    "   Puste pole = pozycja pojawi się w przedmiarze bez ceny i bez wartości.",
    "4. SAMOROZBUDOWA: gdy program zyska nową funkcję generującą nową pozycję",
    "   przedmiaru (nowy element hali), przy najbliższym liczeniu przedmiaru wiersz",
    "   dla tej pozycji zostanie automatycznie dopisany na końcu arkusza z pustą ceną —",
    "   wystarczy wtedy uzupełnić cenę ręcznie. Nic nie trzeba dopisywać samodzielnie.",
    "5. Po edycji zapisz plik i wgraj go z powrotem przyciskiem 'Wgraj cennik' w aplikacji",
    "   (przycisk 'Pobierz cennik' służy do pobrania aktualnej wersji do edycji).",
    "   Wgranie scala Twoje zmiany z serwerem — pozycje dopisane w międzyczasie przez",
    "   innych użytkowników nie zostaną skasowane.",
    "6. Nie zostawiaj pustych wierszy pomiędzy danymi w arkuszu 'Ceny' — pusty",
    "   'Opis pozycji' kończy odczyt arkusza.",
    "7. Ceny są NETTO, w PLN — sposób ich interpretacji (netto/brutto) ustala użytkownik",
    "   przy analizie wyniku przedmiaru.",
]

_LOCK = threading.Lock()


class PriceCatalogValidationError(ValueError):
    """Przesłany plik nie jest poprawnym katalogiem cen (zły arkusz/kolumny/pusty)."""


# --- Odczyt ---

def _read_price_rows(path=None) -> List[dict]:
    wb = load_workbook(path or CATALOG_PATH, read_only=True, data_only=True)
    try:
        ws = wb[PRICES_SHEET]
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        idx = {}
        for name in PRICE_COLUMNS:
            if name not in header:
                raise ValueError(f"Brak kolumny '{name}' w arkuszu '{PRICES_SHEET}'")
            idx[name] = header.index(name)

        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            opis = row[idx[COL_OPIS]] if idx[COL_OPIS] < len(row) else None
            if opis is None or str(opis).strip() == "":
                continue
            rows.append({col: (row[idx[col]] if idx[col] < len(row) else None) for col in PRICE_COLUMNS})
        return rows
    finally:
        wb.close()


def load_price_catalog() -> List[dict]:
    """
    Zwraca listę pozycji cennika (np. do podglądu w UI).
    Jeśli plik jest w danej chwili niedostępny do odczytu albo uszkodzony —
    zwraca listę pustą zamiast wyjątku, żeby liczenie przedmiaru mogło się
    mimo to zakończyć (po prostu bez cen w tym przebiegu).
    """
    if not os.path.exists(CATALOG_PATH):
        return []
    try:
        return _read_price_rows()
    except Exception:
        return []


def _price_map() -> Dict[str, Optional[float]]:
    prices: Dict[str, Optional[float]] = {}
    for row in load_price_catalog():
        opis = str(row[COL_OPIS]).strip()
        cena = row[COL_CENA]
        prices[opis] = float(cena) if isinstance(cena, (int, float)) else None
    return prices


# --- Budowa / zapis pliku (wspólne dla samorozbudowy i wgrywania) ---

def _build_workbook(rows: List[Tuple[str, str, Optional[float], str]]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = PRICES_SHEET

    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(PRICE_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font, cell.fill, cell.alignment, cell.border = header_font, header_fill, center, border

    for r, (opis, jednostka, cena, uwagi) in enumerate(rows, start=2):
        vals = [opis, jednostka, cena, uwagi]
        aligns = [left, center, center, left]
        for c, (val, al) in enumerate(zip(vals, aligns), start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = al
            cell.border = border

    for i, w in enumerate([46, 12, 20, 30], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    instr = wb.create_sheet("Instrukcja")
    for i, line in enumerate(INSTRUCTION_LINES, start=1):
        cell = instr.cell(row=i, column=1, value=line)
        if i == 1:
            cell.font = Font(bold=True, size=12)
    instr.column_dimensions["A"].width = 100

    return wb


def _replace_file(tmp_path: str) -> None:
    """
    Podmienia CATALOG_PATH zawartością tmp_path, trzema metodami o malejącej
    "czystości", każda kolejna próbowana tylko gdy poprzednia zawiedzie:

    1. `os.replace` (MoveFileEx z REPLACE_EXISTING) — atomowe, ale niezbyt
       niezawodne na wolumenach FAT/FAT32.
    2. usuń-i-zmień-nazwę — działa na FAT/FAT32, ale wymaga uprawnienia do
       USUNIĘCIA istniejącego pliku docelowego.
    3. bezpośredni zapis w miejscu (otwarcie CATALOG_PATH do zapisu i
       nadpisanie zawartością tmp_path) — nieatomowe, ale jedyne, co działa,
       gdy zewnętrzny proces (typowo: skanowanie antywirusowe/indeksowanie
       Windows świeżo zmienionego pliku Office) trzyma na pliku blokadę
       odrzucającą usunięcie/zmianę nazwy, a mimo to pozwala go otworzyć do
       zwykłego zapisu — taki stan bywa obserwowany na wolnych dyskach
       wymiennych i może się utrzymywać nawet kilkadziesiąt sekund, więc
       ta metoda jest ostatnią deską ratunku, żeby użytkownik nie dostawał
       błędu mimo że zapis fizycznie jest możliwy.
    """
    try:
        os.replace(tmp_path, CATALOG_PATH)
        return
    except OSError:
        pass

    try:
        os.remove(CATALOG_PATH)
        os.rename(tmp_path, CATALOG_PATH)
        return
    except OSError:
        pass

    with open(tmp_path, "rb") as fh:
        data = fh.read()
    with open(CATALOG_PATH, "wb") as fh:
        fh.write(data)
    os.remove(tmp_path)


def _atomic_save(wb: Workbook, attempts: int = 3, max_delay: float = 0.5) -> None:
    """
    Zapisuje workbook do pliku tymczasowego, po czym podmienia CATALOG_PATH.

    Zarówno `os.replace`, jak i usuń-i-zmień-nazwę bywają na Windows chwilowo
    blokowane przez inny proces (indeksowanie plików, antywirus skanujący
    świeżo zapisane pliki Office) tuż po zapisaniu pliku — mimo że zwykły
    odczyt/zapis samego pliku docelowego działa bez problemu w tym samym
    momencie. Taka blokada bywa krótkotrwała, więc próbę ponawiamy z rosnącym
    opóźnieniem, zanim zgłosimy błąd na dobre.

    `attempts`/`max_delay` pozwalają dopasować cierpliwość do kontekstu:
    krótką dla cichej samorozbudowy w tle (nie może blokować liczenia
    przedmiaru na wiele sekund) i dłuższą dla świadomego wgrywania pliku przez
    użytkownika (tam liczy się skuteczność bardziej niż czas odpowiedzi).
    """
    os.makedirs(_DATA_DIR, exist_ok=True)
    fd, tmp_path = tempfile.mkstemp(suffix=".xlsx", dir=_DATA_DIR)
    os.close(fd)
    try:
        wb.save(tmp_path)
        last_err = None
        for attempt in range(attempts):
            try:
                _replace_file(tmp_path)
                last_err = None
                break
            except OSError as e:
                last_err = e
                if attempt < attempts - 1:
                    time.sleep(min(0.2 * (attempt + 1), max_delay))
        if last_err is not None:
            raise last_err
    except Exception:
        if os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except OSError:
                pass
        raise


def _rows_from_catalog() -> Dict[str, Tuple[str, Optional[float], str]]:
    """opis -> (jednostka, cena, uwagi) na podstawie aktualnego pliku na serwerze."""
    out = {}
    for r in load_price_catalog():
        opis = str(r[COL_OPIS]).strip()
        cena = r[COL_CENA] if isinstance(r[COL_CENA], (int, float)) else None
        out[opis] = (r[COL_JEDNOSTKA] or "", cena, r[COL_UWAGI] or "")
    return out


def _append_missing_rows(missing: List[dict]) -> bool:
    """
    Dopisuje brakujące pozycje (opis + jednostka, bez ceny) do katalogu na serwerze
    (samorozbudowa). Zwraca False po cichu w razie jakiegokolwiek problemu z
    zapisem — dopisanie zostaje odłożone do następnego wywołania.
    """
    try:
        with _LOCK:
            current = _rows_from_catalog()
            for item in missing:
                if item["opis"] not in current:
                    current[item["opis"]] = (item["jednostka"], None, "")
            rows = [(opis, jm, cena, uw) for opis, (jm, cena, uw) in current.items()]
            _atomic_save(_build_workbook(rows))
        return True
    except Exception:
        return False


def sync_and_price_items(items: List[dict]) -> List[dict]:
    """
    Dopasowuje ceny jednostkowe z katalogu do pozycji przedmiaru i dolicza wartość.

    1) Dla pozycji z `items`, których opisu nie ma jeszcze w katalogu — dopisuje
       je do pliku Excel z pustą ceną (samorozbudowa katalogu).
    2) Uzupełnia każdej pozycji cena_jedn (z katalogu, jeśli ustawiona) i wylicza
       wartosc = ilosc * cena_jedn (None, gdy ceny brak).

    Modyfikuje i zwraca tę samą listę `items` (in place), zgodnie z formatem
    zwracanym przez TakeoffCalculator.compute().

    Awaria katalogu (plik uszkodzony, brak arkusza, błąd zapisu) nigdy nie
    przerywa liczenia przedmiaru — w najgorszym razie pozycje zostają bez cen.
    """
    if not os.path.exists(CATALOG_PATH):
        return items

    try:
        existing_opisy = {str(r[COL_OPIS]).strip() for r in load_price_catalog()}

        missing_unique: List[dict] = []
        seen = set()
        for it in items:
            opis = it["opis"]
            if opis not in existing_opisy and opis not in seen:
                seen.add(opis)
                missing_unique.append({"opis": opis, "jednostka": it["jednostka"]})

        if missing_unique:
            _append_missing_rows(missing_unique)

        prices = _price_map()
    except Exception:
        prices = {}

    for it in items:
        cena = prices.get(it["opis"])
        it["cena_jedn"] = cena
        it["wartosc"] = round(it["ilosc"] * cena, 2) if cena is not None else None

    return items


# --- Pobieranie / wgrywanie (edycja po stronie klienta) ---

def get_catalog_bytes() -> bytes:
    """Zwraca surową zawartość pliku katalogu cen — do pobrania przez przeglądarkę."""
    if not os.path.exists(CATALOG_PATH):
        raise FileNotFoundError(f"Brak pliku katalogu cen: {CATALOG_PATH}")
    with open(CATALOG_PATH, "rb") as fh:
        return fh.read()


def _parse_uploaded_rows(file_bytes: bytes) -> Dict[str, Tuple[str, Optional[float], str]]:
    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:
        raise PriceCatalogValidationError(f"Nie udało się odczytać przesłanego pliku jako Excel (.xlsx): {e}")

    if PRICES_SHEET not in wb.sheetnames:
        raise PriceCatalogValidationError(f"W przesłanym pliku brakuje arkusza '{PRICES_SHEET}'.")

    ws = wb[PRICES_SHEET]
    header_row = next(ws.iter_rows(min_row=1, max_row=1), None)
    if header_row is None:
        raise PriceCatalogValidationError(f"Arkusz '{PRICES_SHEET}' jest pusty.")
    header = [c.value for c in header_row]

    idx = {}
    for name in PRICE_COLUMNS:
        if name not in header:
            raise PriceCatalogValidationError(f"Brak kolumny '{name}' w arkuszu '{PRICES_SHEET}' przesłanego pliku.")
        idx[name] = header.index(name)

    rows: Dict[str, Tuple[str, Optional[float], str]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        opis = row[idx[COL_OPIS]] if idx[COL_OPIS] < len(row) else None
        if opis is None or str(opis).strip() == "":
            continue
        opis = str(opis).strip()
        cena_raw = row[idx[COL_CENA]] if idx[COL_CENA] < len(row) else None
        cena = float(cena_raw) if isinstance(cena_raw, (int, float)) else None
        jednostka = row[idx[COL_JEDNOSTKA]] if idx[COL_JEDNOSTKA] < len(row) else ""
        uwagi = row[idx[COL_UWAGI]] if idx[COL_UWAGI] < len(row) else ""
        rows[opis] = (jednostka or "", cena, uwagi or "")

    if not rows:
        raise PriceCatalogValidationError(f"Arkusz '{PRICES_SHEET}' przesłanego pliku nie zawiera żadnych pozycji.")
    return rows


def import_price_catalog(file_bytes: bytes) -> dict:
    """
    Wgrywa zedytowany przez użytkownika plik cennika, SCALAJĄC go z aktualnym
    katalogiem na serwerze (nie zastępuje w ciemno):
    - pozycje obecne w przesłanym pliku nadpisują cenę/jednostkę/uwagi na serwerze,
    - pozycje, które są na serwerze, ale nie wystąpiły w przesłanym pliku (np. zostały
      dopisane automatycznie w międzyczasie przez innego użytkownika) są zachowywane,
    - zupełnie nowe pozycje, obecne tylko w przesłanym pliku, również zostają dodane.

    Rzuca PriceCatalogValidationError, jeśli plik nie jest poprawnym katalogiem cen.
    Zapis jest atomowy i zserializowany (_LOCK) względem innych zapisów do katalogu.
    """
    uploaded = _parse_uploaded_rows(file_bytes)

    with _LOCK:
        server_rows = _rows_from_catalog()

        matched = sum(1 for opis in uploaded if opis in server_rows)
        added = len(uploaded) - matched
        kept_only_on_server = len(server_rows) - matched

        merged_order = list(server_rows.keys())
        for opis in uploaded:
            if opis not in server_rows:
                merged_order.append(opis)

        merged = dict(server_rows)
        merged.update(uploaded)

        rows = [(opis, *merged[opis]) for opis in merged_order]
        # Dłuższa cierpliwość niż przy cichej samorozbudowie w tle — to świadome,
        # rzadkie działanie użytkownika, gdzie liczy się skuteczność zapisu bardziej
        # niż czas odpowiedzi (np. Windows Search potrafi na wolnym dysku USB
        # trzymać świeżo zapisany plik Office zablokowany do indeksowania nawet
        # kilkanaście-kilkadziesiąt sekund).
        _atomic_save(_build_workbook(rows), attempts=15, max_delay=2.0)

    return {
        "total": len(rows),
        "updated_existing": matched,
        "added_new": added,
        "kept_from_server_only": kept_only_on_server,
    }
