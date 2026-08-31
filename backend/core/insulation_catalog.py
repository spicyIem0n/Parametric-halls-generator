"""
InsulationCatalog — wczytuje katalog materiałów izolacji dachowej
(termicznej i przeciwwodnej) z pliku Excel data/roof_insulation_catalog.xlsx.

Plik jest edytowalny przez użytkownika — nowe materiały/warianty grubości
dopisuje się jako kolejne wiersze arkuszy "Izolacja_termiczna" i
"Izolacja_przeciwwodna" (patrz arkusz "Instrukcja" w pliku).
Katalog jest wczytywany na żądanie (bez cache), więc zmiany w pliku Excel
są widoczne od razu po odświeżeniu w aplikacji.
"""
import os
from openpyxl import load_workbook

_DATA_DIR = os.path.join(os.path.dirname(__file__), "..", "data")
CATALOG_PATH = os.path.join(_DATA_DIR, "roof_insulation_catalog.xlsx")

THERMAL_SHEET = "Izolacja_termiczna"
WATERPROOFING_SHEET = "Izolacja_przeciwwodna"

THERMAL_COLUMNS = ["ID", "Materiał", "Grubość [cm]", "Lambda [W/mK]", "Ciężar właściwy [kg/m3]"]
WATERPROOFING_COLUMNS = ["ID", "Materiał", "Grubość [mm]", "Ciężar właściwy [kg/m3]"]


def _read_sheet(sheet_name: str, columns: list[str]) -> list[dict]:
    wb = load_workbook(CATALOG_PATH, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name]
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        idx = {}
        for name in columns:
            if name not in header:
                raise ValueError(f"Brak kolumny '{name}' w arkuszu '{sheet_name}' pliku {CATALOG_PATH}")
            idx[name] = header.index(name)

        items = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            id_val = row[idx["ID"]] if idx["ID"] < len(row) else None
            if id_val is None or str(id_val).strip() == "":
                continue
            item = {col: row[idx[col]] if idx[col] < len(row) else None for col in columns}
            items.append(item)
        return items
    finally:
        wb.close()


def load_thermal_insulation_catalog() -> list[dict]:
    """Zwraca listę wariantów izolacji termicznej dachu (materiał + grubość + lambda + ciężar)."""
    return _read_sheet(THERMAL_SHEET, THERMAL_COLUMNS)


def load_waterproofing_catalog() -> list[dict]:
    """Zwraca listę wariantów izolacji przeciwwodnej dachu (materiał + grubość + ciężar)."""
    return _read_sheet(WATERPROOFING_SHEET, WATERPROOFING_COLUMNS)
