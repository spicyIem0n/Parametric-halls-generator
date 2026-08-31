"""
SoilCatalog — wczytuje katalog typowych gruntów (nazwa, qdop orientacyjne,
ciężar objętościowy) z pliku Excel data/soil_catalog.xlsx.

Plik jest edytowalny przez użytkownika — nowy rodzaj gruntu dopisuje się jako
kolejny wiersz arkusza "Grunty" (patrz arkusz "Instrukcja" w pliku).
Katalog jest wczytywany na żądanie (bez cache).

UWAGA: wartości qdop są orientacyjne — służą jedynie jako podpowiedź/wartość
startowa. Rzeczywista nośność gruntu MUSI wynikać z dokumentacji geotechnicznej
konkretnej działki.
"""
import os
from openpyxl import load_workbook

_DATA_DIR = os.path.join(os.path.dirname(__file__), "..", "data")
CATALOG_PATH = os.path.join(_DATA_DIR, "soil_catalog.xlsx")

SOIL_SHEET = "Grunty"
SOIL_COLUMNS = ["ID", "Rodzaj gruntu", "qdop [kPa]", "Ciężar objętościowy [kN/m3]"]


def load_soil_catalog() -> list[dict]:
    wb = load_workbook(CATALOG_PATH, read_only=True, data_only=True)
    try:
        ws = wb[SOIL_SHEET]
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        idx = {}
        for name in SOIL_COLUMNS:
            if name not in header:
                raise ValueError(f"Brak kolumny '{name}' w arkuszu '{SOIL_SHEET}' pliku {CATALOG_PATH}")
            idx[name] = header.index(name)

        items = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            id_val = row[idx["ID"]] if idx["ID"] < len(row) else None
            if id_val is None or str(id_val).strip() == "":
                continue
            item = {col: row[idx[col]] if idx[col] < len(row) else None for col in SOIL_COLUMNS}
            items.append(item)
        return items
    finally:
        wb.close()
