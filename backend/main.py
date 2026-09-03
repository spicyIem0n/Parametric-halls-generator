import io
from fastapi import FastAPI, HTTPException, UploadFile, File, Header
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from models import HallParameters
from generators.hall_generator import HallGenerator
from core.grid_system import GridSystem3D
from core.clash_detector import ClashDetector
from core.takeoff_calculator import TakeoffCalculator
from core.insulation_catalog import load_thermal_insulation_catalog, load_waterproofing_catalog
from core.roof_load_calculator import RoofLoadCalculator
from core.foundation_sizing_calculator import FoundationSizingCalculator
from core.soil_catalog import load_soil_catalog
from core.ifc_exporter import export_ifc_to_bytes
from core.price_catalog import (
    load_price_catalog,
    sync_and_price_items,
    get_catalog_bytes,
    import_price_catalog,
    PriceCatalogValidationError,
)
from core.feature_flags import get_flags, set_flag, is_enabled, FLAG_LABELS, DEFAULT_FLAGS
from core.admin_auth import check_admin_token

app = FastAPI(title="Parametric Hall API")

# Konfiguracja CORS (niezbędne dla połączenia z Reactem)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # W produkcji zamień na domenę frontendu
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.post("/generate-hall")
def generate_hall(params: HallParameters):
    """Generuje pełną geometrię 3D hali na podstawie parametrów."""
    generator = HallGenerator(params)
    components = generator.generate_all_components()
    return {"components": components}


@app.post("/export/ifc")
def export_ifc(params: HallParameters):
    """
    Eksportuje wygenerowany model 3D hali do pliku IFC4 (do otwarcia w Revit, BimVision itp.).

    Faza 1: masing/koordynacja — elementy jako bryły prostopadłościenne
    (IfcColumn/IfcBeam/IfcWall/IfcSlab/... wg mapowania w core/ifc_exporter.py),
    poprawnie rozmieszczone i zorientowane w przestrzeni (konwersja Y-up -> Z-up).
    """
    generator = HallGenerator(params)
    components = generator.generate_all_components()
    ifc_bytes = export_ifc_to_bytes(params, components)

    width = int(params.width) if params.width else 0
    length = int(params.length) if params.length else 0
    fname = f"hala_{width}x{length}.ifc"

    return StreamingResponse(
        io.BytesIO(ifc_bytes),
        media_type="application/x-step",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@app.post("/validate-hall")
def validate_hall(params: HallParameters):
    """
    Waliduje model hali pod kątem kolizji geometrycznych.
    Zwraca listę ostrzeżeń i błędów (clashów).
    """
    if params.hall_type == "complex":
        # Dla hal złożonych — walidacja per blok
        all_clashes = []
        for block in (params.blocks or []):
            block_dict = params.model_dump()
            block_dict.update({
                "hall_type": "simple",
                "width": block.width, "length": block.length,
                "clear_height": block.clear_height, "bay_spacing": block.bay_spacing,
                "roof_angle": block.roof_angle, "roof_drainage_type": block.roof_drainage_type,
                "number_of_aisles": block.number_of_aisles,
                "docks_config": {}, "blocks": [], "fire_walls": [],
                "technical_rooms": [], "external_offices": [],
                "internal_offices": [], "office_reserve_zones": [],
            })
            block_params = HallParameters(**block_dict)
            grid = GridSystem3D(block_params)
            block_params.length = grid.length
            detector = ClashDetector(grid, block_params)
            result = detector.validate()
            for clash in result.clashes:
                clash.message = f"[{block.block_id}] {clash.message}"
            all_clashes.extend(result.clashes)

        errors = [c for c in all_clashes if c.severity == "error"]
        return {
            "is_valid": len(errors) == 0,
            "warnings_count": len(all_clashes) - len(errors),
            "errors_count": len(errors),
            "clashes": [
                {"clash_type": c.clash_type, "severity": c.severity, "message": c.message,
                 "element_a": c.element_a_type, "element_b": c.element_b_type,
                 "position": c.position, "bay_index": c.bay_index, "side": c.side}
                for c in all_clashes
            ],
        }

    grid = GridSystem3D(params)
    params.length = grid.length

    detector = ClashDetector(grid, params)
    result = detector.validate()

    return result.to_dict()


@app.get("/catalogs/roof-thermal-insulation")
def roof_thermal_insulation_catalog():
    """Katalog materiałów izolacji termicznej dachu (wczytywany z pliku Excel)."""
    return {"items": load_thermal_insulation_catalog()}


@app.get("/catalogs/roof-waterproofing")
def roof_waterproofing_catalog():
    """Katalog materiałów izolacji przeciwwodnej dachu (wczytywany z pliku Excel)."""
    return {"items": load_waterproofing_catalog()}


@app.post("/roof-loads")
def roof_loads(params: HallParameters):
    """Zebranie obciążeń dachu (wartości charakterystyczne) per moduł/hala."""
    return RoofLoadCalculator.compute(params)


@app.post("/foundation-sizing")
def foundation_sizing(params: HallParameters):
    """Automatyczny dobór gabarytów stóp fundamentowych (wartości orientacyjne)."""
    return FoundationSizingCalculator.compute(params)


@app.get("/catalogs/soil")
def soil_catalog():
    """Katalog typowych gruntów z orientacyjnym qdop (wczytywany z pliku Excel)."""
    return {"items": load_soil_catalog()}


@app.get("/catalogs/prices")
def price_catalog():
    """Katalog cen jednostkowych pozycji przedmiaru (wczytywany z pliku Excel)."""
    return {"items": load_price_catalog()}


@app.get("/catalogs/prices/download")
def download_price_catalog():
    """
    Pobiera plik katalogu cen (price_catalog.xlsx) na komputer użytkownika —
    do edycji we własnym, lokalnym Excelu. Plik na serwerze pozostaje
    niezmieniony (to tylko odczyt), więc pobieranie nigdy nie koliduje z
    żadnym innym użytkownikiem ani z samorozbudową katalogu.
    """
    if not is_enabled("price_catalog_edit"):
        raise HTTPException(status_code=403, detail="Ta funkcja jest wyłączona w tej wersji programu.")
    try:
        data = get_catalog_bytes()
    except FileNotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="price_catalog.xlsx"'},
    )


@app.post("/catalogs/prices/upload")
async def upload_price_catalog(file: UploadFile = File(...)):
    """
    Wgrywa zedytowany lokalnie plik katalogu cen z powrotem na serwer.
    Zawartość jest SCALANA z aktualnym stanem serwera (pozycje dopisane przez
    innych użytkowników w międzyczasie nie są kasowane) i zapisywana atomowo.
    """
    if not is_enabled("price_catalog_edit"):
        raise HTTPException(status_code=403, detail="Ta funkcja jest wyłączona w tej wersji programu.")
    content = await file.read()
    try:
        summary = import_price_catalog(content)
    except PriceCatalogValidationError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except (PermissionError, OSError):
        raise HTTPException(
            status_code=503,
            detail="Serwer chwilowo nie mógł zapisać pliku katalogu cen (jest w tej chwili używany przez inny "
                   "proces na serwerze). Spróbuj wgrać plik ponownie za chwilę.",
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Nie udało się zapisać katalogu cen: {e}")
    return summary


@app.get("/features")
def features():
    """
    Stan przełączników funkcji programu (widoczny publicznie — frontend na jego
    podstawie pokazuje/ukrywa dane funkcje, np. dla wersji trialowej).
    """
    return {"flags": get_flags()}


@app.post("/admin/features")
def set_feature_flag(payload: dict, x_admin_token: str = Header(default="")):
    """
    Włącza/wyłącza jedną flagę funkcji. Wymaga poprawnego tokenu administratora
    w nagłówku X-Admin-Token (patrz core.admin_auth / zmienna ADMIN_TOKEN).
    Body: {"name": "<nazwa_flagi>", "value": true|false}.
    """
    if not check_admin_token(x_admin_token):
        raise HTTPException(status_code=401, detail="Nieprawidłowy token administratora.")
    name = payload.get("name")
    value = payload.get("value")
    if name not in DEFAULT_FLAGS or not isinstance(value, bool):
        raise HTTPException(
            status_code=400,
            detail=f"Nieprawidłowe dane. Znane flagi: {list(DEFAULT_FLAGS.keys())}.",
        )
    try:
        flags = set_flag(name, value)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Nie udało się zapisać ustawienia: {e}")
    return {"flags": flags}


@app.post("/admin/verify")
def verify_admin_token(x_admin_token: str = Header(default="")):
    """Sprawdza, czy podany token administratora jest poprawny (do logowania w panelu)."""
    if not check_admin_token(x_admin_token):
        raise HTTPException(status_code=401, detail="Nieprawidłowy token administratora.")
    return {"ok": True, "labels": FLAG_LABELS}


@app.post("/quantity-takeoff")
def quantity_takeoff(params: HallParameters):
    """Zwraca przedmiar ilosciowy (lista pozycji) na podstawie modelu, wyceniony wg katalogu cen."""
    items = TakeoffCalculator.compute(params)
    items = sync_and_price_items(items)
    return {"items": items}


@app.post("/quantity-takeoff/export")
def quantity_takeoff_export(params: HallParameters):
    """Eksportuje wyceniony przedmiar do pliku Excel (.xlsx)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    items = TakeoffCalculator.compute(params)
    items = sync_and_price_items(items)

    wb = Workbook()
    ws = wb.active
    ws.title = "Przedmiar"

    headers = ["L.p.", "Opis pozycji", "Jednostka miary", "Ilość",
               "Cena jednostkowa", "Wartość", "Uwagi"]

    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=1 + col - 1, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    for r, it in enumerate(items, start=2):
        row_vals = [
            it["lp"], it["opis"], it["jednostka"], it["ilosc"],
            it["cena_jedn"], it["wartosc"], it["uwagi"],
        ]
        for c, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = border
            if c in (1, 3, 4, 5, 6):
                cell.alignment = center
            else:
                cell.alignment = left

    # Wiersz podsumowania (suma pozycji z wypelniona cena; pozycje bez ceny nie wchodza do sumy)
    total = sum(it["wartosc"] for it in items if it["wartosc"] is not None)
    priced_count = sum(1 for it in items if it["wartosc"] is not None)
    total_row = len(items) + 2
    ws.cell(row=total_row, column=5, value="RAZEM:").font = Font(bold=True)
    ws.cell(row=total_row, column=5).alignment = Alignment(horizontal="right")
    total_cell = ws.cell(row=total_row, column=6, value=round(total, 2))
    total_cell.font = Font(bold=True)
    total_cell.alignment = center
    if priced_count < len(items):
        ws.cell(row=total_row, column=7,
                value=f"Uwaga: {len(items) - priced_count} z {len(items)} pozycji bez ceny w katalogu — suma częściowa.")

    # Szerokosci kolumn
    widths = [6, 42, 14, 12, 16, 14, 26]
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"przedmiar_hala_{int(params.width)}x{int(params.length)}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# Aby uruchomić: uvicorn main:app --reload
